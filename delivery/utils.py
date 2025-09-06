import httplib2
import apiclient.discovery
from oauth2client.service_account import ServiceAccountCredentials
from deliveryplus import settings
from datetime import date
import io
from reportlab.pdfgen import canvas
from transaction.models import Transaction

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# open credential file
CREDENTIALS_FILE = "cred.json"
# ID Google Sheets documents
spreadsheet_id = settings.SPREADSHEET_ID

# service — API access instance
credentials = ServiceAccountCredentials.from_json_keyfile_name(
    CREDENTIALS_FILE,
    [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ],
)
httpAuth = credentials.authorize(httplib2.Http())
service = apiclient.discovery.build("sheets", "v4", http=httpAuth)


def write_report_gs(data=None, sheet_name=None):
    request_body = {"values": data}

    clear_request = (
        service.spreadsheets()
        .values()
        .clear(
            spreadsheetId=spreadsheet_id,
            range=f"{sheet_name}!A1:Z",  # Adjust range as needed
            body={},
        )
    )
    clear_response = clear_request.execute()

    # Call the Sheets API to append the data
    request = (
        service.spreadsheets()
        .values()
        .update(
            spreadsheetId=spreadsheet_id,
            range=f"{sheet_name}!A1",
            valueInputOption="USER_ENTERED",
            body=request_body,
        )
    )
    response = request.execute()

def generate_deliveries_excel(deliveries):
    """Генерує Excel-файл з переданих Delivery записів."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Completed Deliveries"

    headers = [
        "Numer zamówienia", "Identyfikator", "SSCC", "Data otrzymania",
        "Supplier", "Lokalizacja przyjęciowa", "Obecna lokalizacja",
        "Sklep", "Uzytkownik ktory przyjal towar", "Komentarz", "Transaction"
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for d in deliveries:
        if d.transaction:
            lines = [line.strip() for line in d.transaction.splitlines() if line.strip()]
            transaction_text = "\n".join(lines)
        else:
            transaction_text = ""

        row = [
            str(d.nr_order) if d.nr_order else "",
            str(d.identifier) if d.identifier else "",
            str(d.sscc_barcode) if d.sscc_barcode else "",
            str(d.date_recive.strftime("%Y-%m-%d")) if d.date_recive else "",
            str(d.supplier_company.name) if d.supplier_company else "",
            str(d.recive_location.name) if d.recive_location else "",
            str(d.location.name) if d.location else "",
            str(d.shop.name) if d.shop else "",
            str(d.user) if d.user else "",
            str(d.comment.replace("Podcas rozładunku wykryto ", "").replace("Podczas kontroli wykryto ", "")) if d.comment else "",
            transaction_text
        ]
        ws.append(row)

    # Wrap text
    transaction_col = len(headers)
    for cell in ws[get_column_letter(transaction_col)]:
        cell.alignment = Alignment(wrap_text=True)

    # Автоширина
    for i, column_cells in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 40)

    return wb

def create_transaction(user, delivery, transaction_type):
    transaction = Transaction(name=transaction_type, user=user, delivery=delivery)
    transaction.save()

def gen_comment(request):
    index = 0
    reasones = request.POST.get("reasones")
    rec_loc = request.POST.get("recive_location")
    if rec_loc == "second":
        ean_qty_str = ""
        while request.POST.get(f"qty_{index}"):
            qty = request.POST.get(f"qty_{index}")
            ean = request.POST.get(f"ean_{index}", "")
            ean_qty_str += f"{ean} {qty} szt. "
            index += 1
        comment = f"{reasones}: {ean_qty_str}"
    else:
        qty_unit = request.POST.get("qty_unit")
        tape_of_delivery = request.POST.get("tape_of_delivery")
        comment = f"{reasones}: {qty_unit} szt. {tape_of_delivery}"
    return comment

def get_smart_split_comment(comment):
    smart_split_text = []
    line = ""
    for word in comment.split(" "):
        if len(line) + len(word) >= 100:
            smart_split_text.append(line)
            line = f"{word}"
        else:
            line += " " + word
    if line:
        smart_split_text.append(line)
    return smart_split_text

def get_total_products_count(comment):
    count = 0
    split_comment = comment.split(" ")[::-1]
    for indx, word in enumerate(split_comment):
        if "szt" in word.lower():
            try:
                count += int(split_comment[indx + 1])
            except:
                pass
    return count


def gen_pdf_damage_repor(delivery):
    recive_loc = delivery.recive_location.name
    order = delivery.nr_order
    shop = delivery.shop
    supplier = delivery.supplier_company.name
    full_name = delivery.user.full_name
    recive_data = delivery.date_recive.strftime("%Y-%m-%d")
    # full_comment = comment + extra_comment
    comment = delivery.comment
    total_qty = get_total_products_count(comment)
    extra_commrnt = delivery.extra_comment
    sscc = delivery.sscc_barcode

    if shop is not None:
        shop = shop.position_nr
    else:
        shop = ""

    if recive_loc == "1R":
        damage_protocol_path = "static/img/first_rec.jpg"
    else:
        damage_protocol_path = "static/img/second_rec.jpg"

    buffer = io.BytesIO()
    pdfmetrics.registerFont(TTFont("FreeSans", "freesans/FreeSans.ttf"))
    my_canvas = canvas.Canvas(buffer)
    my_canvas.drawImage(damage_protocol_path, 0, 0, width=602, height=840)
    my_canvas.setFont("FreeSans", 10)

    if recive_loc == "1R":
        my_canvas.drawString(170, 664, f"{full_name}")
        my_canvas.drawString(342, 664, f"{recive_data}")
        my_canvas.drawString(50, 507, f"{order}")
        if shop:
            my_canvas.drawString(140, 507, f"LM - {shop}")
        my_canvas.drawString(210, 507, f"{total_qty} {'Paczka' if 'pacz.' in comment else 'Paleta'}.")
        my_canvas.drawString(320, 507, f"{supplier}")
        #my_canvas.drawString(210, 500, f"{'Paczka' if 'pacz.' in comment else 'Paleta'}")

        line_spacing = 20

        x_position = 40
        y_position = 227

        my_canvas.drawString(x_position, 245, f"SSCC/ASN: {sscc}")
        for line in get_smart_split_comment(comment=comment.replace("pacz.", "").replace("pall.", "")):
            my_canvas.drawString(x_position, y_position, f"{line}")
            y_position -= line_spacing
        if extra_commrnt:
            for line in get_smart_split_comment(comment=extra_commrnt):
                my_canvas.drawString(x_position, y_position, f"{line}.")
                y_position -= line_spacing
    else:
        my_canvas.drawString(170, 666, f"{full_name}")
        my_canvas.drawString(342, 666, f"{recive_data}")
        my_canvas.drawString(50, 507, f"{order}")
        my_canvas.drawString(140, 507, f"LM - {shop}")
        my_canvas.drawString(210, 507, f"{total_qty} szt.")
        my_canvas.drawString(300, 507, f"{supplier[:16]}")

        line_spacing = 20
        x_position = 40
        y_position = 232

        my_canvas.drawString(x_position, 250, f"SSCC: {sscc}")
        for line in get_smart_split_comment(comment=comment):
            my_canvas.drawString(x_position, y_position, f"{line}")
            y_position -= line_spacing
        if extra_commrnt:
            my_canvas.drawString(x_position, y_position, f"{extra_commrnt}.")

    my_canvas.showPage()
    my_canvas.save()
    buffer.seek(0)
    return buffer


if __name__ == "main":
    pass
